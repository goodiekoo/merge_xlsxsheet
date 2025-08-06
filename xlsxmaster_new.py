import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import time
from copy import copy
import warnings
from concurrent.futures import ThreadPoolExecutor
import threading

def check_xlsx_sheets(directory_path):
    print(f"\n=== {directory_path} xlsx Infos ===")
    
    xlsx_files = []
    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                xlsx_files.append(os.path.join(root, file))
    
    if not xlsx_files:
        print("Excel does not exist")
        return {}
    
    print(f"Finds: {len(xlsx_files)}")
    file_info = {}
    
    for i, file_path in enumerate(xlsx_files, 1):
        try:
            if len(xlsx_files) > 10 and i % 10 == 0:
                print(f"  Searching: {i}/{len(xlsx_files)} ...")
            
            # openpyxl
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            sheet_count = len(sheet_names)
            
            file_name = os.path.basename(file_path)
            file_info[file_name] = {
                'path': file_path,
                'sheet_count': sheet_count,
                'sheet_names': sheet_names
            }

            if len(xlsx_files) <= 20:
                print(f"\nFile Name: {file_name}")
                print(f"Sheet Count: {sheet_count}")
                print(f"Sheet Name: {sheet_names}")
                
                for sheet_name in sheet_names:
                    try:
                        # openpyx
                        worksheet = workbook[sheet_name]
                        max_row = worksheet.max_row if worksheet.max_row else 0
                        max_col = worksheet.max_column if worksheet.max_column else 0

                        print(f"  - {sheet_name}: {max_row}row × {max_col}col")
                    except Exception as e:
                        print(f"  - {sheet_name}: Error! - {str(e)}")
            else:
                # 파일이 많을 때는 요약 정보만
                if i <= 5:  # 처음 5개만 자세히 표시
                    print(f"\nFile: {file_name} (Sheet: {sheet_count}개)")
                elif i == 6:
                    print(f"\n... (Summary) ...")
            
            workbook.close()
            
        except Exception as e:
            print(f"File {os.path.basename(file_path)} read error: {str(e)}")

    print(f"\n Result: {len(file_info)} files collected.")
    return file_info

def copy_sheet_with_styles(src_sheet, dest_sheet):
    """
    Keep the cell styles when copying a sheet.
    """
    dest_sheet.sheet_view.selection[0].activeCell = src_sheet.sheet_view.selection[0].activeCell
    dest_sheet.sheet_view.selection[0].sqref = src_sheet.sheet_view.selection[0].sqref

    for col in range(1, src_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        dest_sheet.column_dimensions[col_letter].width = src_sheet.column_dimensions[col_letter].width
    for row in range(1, src_sheet.max_row + 1):
        dest_sheet.row_dimensions[row].height = src_sheet.row_dimensions[row].height

    for row in src_sheet.iter_rows():
        for cell in row:
            new_cell = dest_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

def process_file(file_path, key_prefix, output_workbook, lock, processed_sheets_count):
    try:
        # print(f"Processing: {os.path.basename(file_path)}")
        src_workbook = load_workbook(file_path, data_only=True)
        count_sheets = len(src_workbook.sheetnames)
        # print(f"  - Sheet count: {count_sheets}")
        
        for sheet_name in src_workbook.sheetnames:
            src_sheet = src_workbook[sheet_name]
            korean_found = False
            for row in src_sheet.iter_rows():
                row_values = [str(cell.value).lower() if cell.value is not None else '' for cell in row]
                # 'Korean' 또는 'korean'이 포함된 셀을 대소문자 구분 없이 찾음
                if any(cell_val in ["korean", "korean"] for cell_val in row_values):
                    korean_found = True
            
            if korean_found:
                file_name_no_ext = os.path.splitext(os.path.basename(file_path))[0]
                
                new_sheet_name = f"{file_name_no_ext}@{count_sheets}@{sheet_name}"
                safe_sheet_name = new_sheet_name[:31]
                
                with lock:
                    dest_sheet = output_workbook.create_sheet(title=safe_sheet_name)
                    copy_sheet_with_styles(src_sheet, dest_sheet)
                    processed_sheets_count.value += 1
        
        src_workbook.close()
    except Exception as e:
        print(f"    ERROR: {os.path.basename(file_path)} - {str(e)}")

def process_and_merge_files(file_paths, output_path, key_prefix=""):
    """
    지정된 파일 목록을 처리하고 유효한 시트를 병합하여 Excel 파일로 저장합니다. (멀티스레딩 적용)
    """
    folder_name = "Script" if key_prefix else "Ingame"
    if not file_paths:
        print(f"{folder_name} has no files to process.")
        return

    print(f"\n--- {folder_name} files processing and merging ---")

    output_workbook = openpyxl.Workbook()
    if 'Sheet' in output_workbook.sheetnames:
        output_workbook.remove(output_workbook['Sheet'])

    lock = threading.Lock()
    
    # Using a mutable object to count processed sheets across threads
    class Counter:
        def __init__(self):
            self.value = 0

    processed_sheets_counter = Counter()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(process_file, file_path, key_prefix, output_workbook, lock, processed_sheets_counter) for file_path in file_paths]
            for future in futures:
                future.result() # Wait for all threads to complete

    if processed_sheets_counter.value > 0:
        print(f"\n--- Saving merged file: {output_path} ---")
        print(f"Total {processed_sheets_counter.value} valid sheets will be saved...")
        output_workbook.save(output_path)
        print(f"Merge complete! Saved to {output_path}.")
    else:
        print(f"\n{folder_name} has no valid sheets to merge.")
    output_workbook.close()

def merge_xlsx_files(ingame_path, script_path, base_output_path):
    """
    In order to merge Excel files from ingame and script folders into separate files.
    """
    print("\n=== Prepare Merging ===")

    ingame_files = []
    if os.path.exists(ingame_path):
        for root, _, files in os.walk(ingame_path):
            for file in files:
                if file.endswith('.xlsx') and not file.startswith('~$'):
                    ingame_files.append(os.path.join(root, file))
    
    script_files = []
    if os.path.exists(script_path):
        for root, _, files in os.walk(script_path):
            for file in files:
                if file.endswith('.xlsx') and not file.startswith('~$'):
                    script_files.append(os.path.join(root, file))

    print(f"Ingame folder found files: {len(ingame_files)}")
    print(f"Script folder found files: {len(script_files)}")

    timestamp = time.strftime("%m%d")
    ingame_output_path = os.path.join(base_output_path, f"ingame_{timestamp}.xlsx")
    process_and_merge_files(ingame_files, ingame_output_path, key_prefix="")

    script_output_path = os.path.join(base_output_path, f"script_{timestamp}.xlsx")
    process_and_merge_files(script_files, script_output_path, key_prefix="[script]")

def main():
    start_time = time.time()

    base_path = r"...\localization_text"
    ingame_path = os.path.join(base_path, "ingame")
    script_path = os.path.join(base_path, "script")
    # 결과 파일을 'output' 폴더에 저장하도록 변경
    output_base_path = os.path.join(base_path, "output")
    
    print("=== Exceptions ===")
    print(f"Base Dir: {base_path}")
    
    if not os.path.exists(base_path):
        print(f"Error doesn't exist: {base_path}")
        return
        
    if not os.path.exists(output_base_path):
        os.makedirs(output_base_path)
        print(f"Output directory created: {output_base_path}")

    # The check_xlsx_sheets calls are not strictly necessary for merging
    # and can be commented out to speed up the process if not needed.
    # check_xlsx_sheets(base_path)
    
    ingame_exists = os.path.exists(ingame_path)
    if not ingame_exists:
        print(f"\nWarning: Ingame folder doesn't exist: {ingame_path}")

    script_exists = os.path.exists(script_path)
    if not script_exists:
        print(f"\nWarning: Script folder doesn't exist: {script_path}")

    if ingame_exists or script_exists:
        merge_xlsx_files(ingame_path, script_path, output_base_path)
    else:
        print("\nNo folders found to merge Excel files.")

    end_time = time.time()
    execution_time = end_time - start_time

    print(f"\n=== Finish ===")
    print(f"Total execution time: {execution_time:.2f}s")

if __name__ == "__main__":
    main()
